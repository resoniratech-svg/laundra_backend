from sqlalchemy.orm import Session
from uuid import UUID
from datetime import date, datetime
import csv
import io
from openpyxl import Workbook
from fpdf import FPDF

from app.repositories.report_repository import ReportRepository

class ReportService:
    @staticmethod
    def get_dashboard_summary(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_dashboard_summary(db, tenant_id)

    @staticmethod
    def get_daily_sales(db: Session, tenant_id: UUID, target_date: date) -> dict:
        return ReportRepository.get_daily_sales(db, tenant_id, target_date)

    @staticmethod
    def get_monthly_sales(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_monthly_sales(db, tenant_id)

    @staticmethod
    def get_yearly_sales(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_yearly_sales(db, tenant_id)

    @staticmethod
    def get_orders_by_status(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_orders_by_status(db, tenant_id)

    @staticmethod
    def get_orders_between_dates(db: Session, tenant_id: UUID, start: date, end: date) -> list:
        return ReportRepository.get_orders_between_dates(db, tenant_id, start, end)

    @staticmethod
    def get_top_customers(db: Session, tenant_id: UUID, sort_by: str) -> list:
        return ReportRepository.get_top_customers(db, tenant_id, sort_by)

    @staticmethod
    def get_customer_growth(db: Session, tenant_id: UUID) -> list:
        return ReportRepository.get_customer_growth(db, tenant_id)

    @staticmethod
    def get_customer_wallet_report(db: Session, tenant_id: UUID) -> list:
        return ReportRepository.get_customer_wallet_report(db, tenant_id)

    @staticmethod
    def get_revenue_by_payment_method(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_revenue_by_payment_method(db, tenant_id)

    @staticmethod
    def get_pending_payments(db: Session, tenant_id: UUID) -> list:
        return ReportRepository.get_pending_payments(db, tenant_id)

    @staticmethod
    def get_failed_payments(db: Session, tenant_id: UUID) -> list:
        return ReportRepository.get_failed_payments(db, tenant_id)

    @staticmethod
    def get_delivery_performance(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_delivery_performance(db, tenant_id)

    @staticmethod
    def get_delivery_boy_performance(db: Session, tenant_id: UUID) -> list:
        return ReportRepository.get_delivery_boy_performance(db, tenant_id)

    @staticmethod
    def get_expenses_report(db: Session, tenant_id: UUID, category: str) -> list:
        return ReportRepository.get_expenses_report(db, tenant_id, category)

    @staticmethod
    def get_monthly_expenses(db: Session, tenant_id: UUID) -> list:
        return ReportRepository.get_monthly_expenses(db, tenant_id)

    @staticmethod
    def get_profit_report(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_profit_report(db, tenant_id)

    @staticmethod
    def get_coupon_report(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_coupon_report(db, tenant_id)

    @staticmethod
    def get_services_report(db: Session, tenant_id: UUID) -> dict:
        return ReportRepository.get_services_report(db, tenant_id)

    # Export Handlers
    @staticmethod
    def export_csv(db: Session, tenant_id: UUID) -> str:
        profit_data = ReportRepository.get_profit_report(db, tenant_id)
        dashboard_data = ReportRepository.get_dashboard_summary(db, tenant_id)
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Laundry Business Performance Summary"])
        writer.writerow([])
        writer.writerow(["Metrics", "Value"])
        writer.writerow(["Today's Orders", dashboard_data["today_orders"]])
        writer.writerow(["Pending Orders", dashboard_data["pending_orders"]])
        writer.writerow(["Completed Orders", dashboard_data["completed_orders"]])
        writer.writerow(["Active Customers", dashboard_data["active_customers"]])
        writer.writerow(["Pending Deliveries", dashboard_data["delivery_pending"]])
        writer.writerow([])
        writer.writerow(["Financial Summary"])
        writer.writerow(["Total Revenue", profit_data["revenue"]])
        writer.writerow(["Total Expenses", profit_data["expenses"]])
        writer.writerow(["Net Profit", profit_data["profit"]])
        
        return output.getvalue()

    @staticmethod
    def export_excel(db: Session, tenant_id: UUID) -> bytes:
        profit_data = ReportRepository.get_profit_report(db, tenant_id)
        dashboard_data = ReportRepository.get_dashboard_summary(db, tenant_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Performance Report"
        
        ws.append(["Laundry Business Performance Summary"])
        ws.append([])
        ws.append(["Metrics", "Value"])
        ws.append(["Today's Orders", dashboard_data["today_orders"]])
        ws.append(["Pending Orders", dashboard_data["pending_orders"]])
        ws.append(["Completed Orders", dashboard_data["completed_orders"]])
        ws.append(["Active Customers", dashboard_data["active_customers"]])
        ws.append(["Pending Deliveries", dashboard_data["delivery_pending"]])
        ws.append([])
        ws.append(["Financial Summary"])
        ws.append(["Total Revenue", float(profit_data["revenue"])])
        ws.append(["Total Expenses", float(profit_data["expenses"])])
        ws.append(["Net Profit", float(profit_data["profit"])])
        
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def export_pdf(db: Session, tenant_id: UUID) -> bytes:
        profit_data = ReportRepository.get_profit_report(db, tenant_id)
        dashboard_data = ReportRepository.get_dashboard_summary(db, tenant_id)
        
        pdf = FPDF()
        pdf.add_page()
        
        # Title
        pdf.set_font("helvetica", "B", 18)
        pdf.cell(0, 15, text="Laundry Business Performance Report", ln=True, align="C")
        pdf.ln(5)
        
        # Operational Summary
        pdf.set_font("helvetica", "B", 14)
        pdf.cell(0, 10, text="Operational Statistics", ln=True)
        pdf.set_font("helvetica", "", 12)
        pdf.cell(100, 8, text=f"Today's Orders: {dashboard_data['today_orders']}", ln=True)
        pdf.cell(100, 8, text=f"Pending Orders: {dashboard_data['pending_orders']}", ln=True)
        pdf.cell(100, 8, text=f"Completed Orders: {dashboard_data['completed_orders']}", ln=True)
        pdf.cell(100, 8, text=f"Active Customers: {dashboard_data['active_customers']}", ln=True)
        pdf.cell(100, 8, text=f"Pending Deliveries: {dashboard_data['delivery_pending']}", ln=True)
        pdf.ln(10)
        
        # Financial Summary
        pdf.set_font("helvetica", "B", 14)
        pdf.cell(0, 10, text="Financial Summary", ln=True)
        pdf.set_font("helvetica", "", 12)
        pdf.cell(100, 8, text=f"Total Revenue: {profit_data['revenue']}", ln=True)
        pdf.cell(100, 8, text=f"Total Expenses: {profit_data['expenses']}", ln=True)
        pdf.cell(100, 8, text=f"Net Profit: {profit_data['profit']}", ln=True)
        
        # Return PDF bytes
        # fpdf2 output(dest="S") returns bytes/string based on version, let's cast or write to BytesIO if needed
        # Actually in fpdf2, pdf.output() without args returns bytearray/bytes.
        return bytes(pdf.output())
